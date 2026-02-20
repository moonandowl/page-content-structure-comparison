"""
Page Structure Comparison Tool - Excel Output Builder
Generates multi-tab Excel workbook with analysis results.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.resolve()
OUTPUTS_DIR = PROJECT_ROOT / "outputs"

# Cell styling
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

# Diagnosis colors
DIAGNOSIS_FILLS = {
    "Content Gap": BLUE_FILL,
    "Authority Gap": ORANGE_FILL,
    "Both": RED_FILL,
    "Competitive": GREEN_FILL,
}


def _page_label(page: dict) -> str:
    """e.g. Dallas #1"""
    return f"{page.get('city', '')} #{page.get('position', '')}"


def _cell_value(ce: dict, key: str) -> str:
    """Convert content element to display value: ✅ Present / ❌ Absent / ⚠️ Partial"""
    el = ce.get(key, {}) if isinstance(ce, dict) else {}
    if isinstance(el, dict):
        present = el.get("present", False)
        if present:
            return "✅ Present"
        # Check for partial
        if el.get("exact_text") or el.get("found") or el.get("claims"):
            return "⚠️ Partial"
    return "❌ Absent"


def _cell_fill(val: str) -> Optional[PatternFill]:
    if "✅" in val:
        return GREEN_FILL
    if "❌" in val:
        return RED_FILL
    if "⚠️" in val:
        return YELLOW_FILL
    return None


# Content element keys in display order
CONTENT_KEYS = [
    "cta_buttons", "video_embed", "faq_section", "testimonials",
    "cost_pricing", "candidacy_quiz", "before_after_photos", "surgeon_credentials",
    "technology_names", "financing", "outcome_statistics", "trust_badges",
    "press_mentions", "live_chat", "online_scheduling", "google_review_widget",
    "video_testimonials",
]

ELEMENT_NAMES = {
    "cta_buttons": "CTA Buttons",
    "video_embed": "Video Embed",
    "faq_section": "FAQ Section",
    "testimonials": "Testimonials/Reviews",
    "cost_pricing": "Cost/Pricing",
    "candidacy_quiz": "Candidacy Quiz",
    "before_after_photos": "Before/After Photos",
    "surgeon_credentials": "Surgeon Credentials",
    "technology_names": "Technology Names",
    "financing": "Financing",
    "outcome_statistics": "Outcome Statistics",
    "trust_badges": "Trust Badges",
    "press_mentions": "Press Mentions",
    "live_chat": "Live Chat",
    "online_scheduling": "Online Scheduling",
    "google_review_widget": "Google Review Widget",
    "video_testimonials": "Video Testimonials",
}


def _qualifying_pages(pages: list[dict]) -> list[dict]:
    """Service Page and Procedure+Location only for content matrix."""
    return [p for p in pages if p.get("page_type") in ("Service Page", "Procedure+Location")]


def build_excel(
    procedure: str,
    merged_data: dict,
    analysis: dict,
    config: dict,
    run_id: Optional[str] = None,
) -> Path:
    """Build Excel workbook. Returns path to saved file."""
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    suffix = f"_{run_id}" if run_id else ""
    filename = f"{procedure}_competitive_analysis_{date_str}{suffix}.xlsx"
    output_path = OUTPUTS_DIR / filename

    wb = Workbook()
    pages = merged_data.get("pages", [])
    qualifying = _qualifying_pages(pages) or pages  # fallback if no Service/Procedure+Location
    coverage = {c["key"]: c for c in analysis.get("content_coverage", [])}

    # --- Tab 1: Master Content Matrix ---
    ws1 = wb.active
    ws1.title = "Master Content Matrix"
    headers = ["Content Element", "Count", "%", "Wireframe Priority"]
    q_labels = [f"{_page_label(p)} ({p.get('page_type', '')})" for p in qualifying]
    headers.extend(q_labels)
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    ws1.freeze_panes = "B2"

    for r, key in enumerate(CONTENT_KEYS, 2):
        cov = coverage.get(key, {})
        ws1.cell(row=r, column=1, value=ELEMENT_NAMES.get(key, key))
        ws1.cell(row=r, column=2, value=cov.get("count", ""))
        ws1.cell(row=r, column=3, value=f"{cov.get('percentage', 0)}%")
        ws1.cell(row=r, column=4, value=cov.get("wireframe_priority", ""))
        if cov.get("position_1_differentiator"):
            ws1.cell(row=r, column=1).fill = GOLD_FILL
        for c, p in enumerate(qualifying, 5):
            ce = p.get("content_elements", {})
            val = _cell_value(ce, key)
            cell = ws1.cell(row=r, column=c, value=val)
            f = _cell_fill(val)
            if f:
                cell.fill = f

    # --- Tab 2: Section Order Map ---
    ws2 = wb.create_sheet("Section Order Map")
    section_order = analysis.get("section_order", {})
    q_labels = [_page_label(p) for p in qualifying]
    consensus = analysis.get("consensus_order", [])
    ws2.cell(row=1, column=1, value="Position")
    for c, lbl in enumerate(q_labels, 2):
        ws2.cell(row=1, column=c, value=lbl).font = Font(bold=True)
    ws2.cell(row=1, column=len(q_labels) + 2, value="Consensus Order").font = Font(bold=True)
    for i, pos in enumerate(sorted(section_order.keys()), 2):
        ws2.cell(row=i, column=1, value=f"{pos}st H2")
        row_data = section_order[pos]
        vals = list(row_data.values())
        for c, lbl in enumerate(q_labels, 2):
            val = row_data.get(lbl, "")
            cell = ws2.cell(row=i, column=c, value=val)
            if vals.count(val) >= 3 or (val and sum(1 for v in vals if v == val) >= 3):
                cell.fill = LIGHT_BLUE_FILL
        if pos - 1 < len(consensus):
            ws2.cell(row=i, column=len(q_labels) + 2, value=consensus[pos - 1])
    ws2.freeze_panes = "B2"

    # --- Tab 2b: Section Word Counts ---
    ws_swc = wb.create_sheet("Section Word Counts")
    swc_headers = ["Page", "Position", "H2 Text", "Word Count"]
    for c, h in enumerate(swc_headers, 1):
        ws_swc.cell(row=1, column=c, value=h).font = Font(bold=True)
    row_num = 2
    for p in qualifying:
        page_label = _page_label(p)
        for sec in p.get("section_word_counts", []):
            ws_swc.cell(row=row_num, column=1, value=page_label)
            ws_swc.cell(row=row_num, column=2, value=sec.get("position", ""))
            ws_swc.cell(row=row_num, column=3, value=sec.get("h2_text", ""))
            ws_swc.cell(row=row_num, column=4, value=sec.get("word_count", 0))
            row_num += 1
    ws_swc.freeze_panes = "B2"

    # --- Tab 3: Authority Profile ---
    ws3 = wb.create_sheet("Authority Profile")
    auth_headers = [
        "City", "Position", "URL", "Page Type", "Classification Confidence",
        "Domain Rating", "URL Rating", "Referring Domains", "Backlinks",
        "Organic Traffic", "Content Richness Score", "Diagnosis", "Notes",
    ]
    for c, h in enumerate(auth_headers, 1):
        ws3.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, p in enumerate(pages, 2):
        notes = []
        if p.get("page_type") == "Homepage":
            notes.append("Homepage ranking — authority driven")
        if p.get("page_type") == "Geo Page":
            notes.append("Geo page ranking — different intent")
        if p.get("js_rendering_flagged"):
            notes.append("⚠️ May require JS rendering")
        notes_str = " ".join(notes)
        row_data = [
            p.get("city", ""),
            p.get("position", ""),
            p.get("url", "")[:80],
            p.get("page_type", ""),
            p.get("page_type_confidence", ""),
            str(p.get("domain_rating", "")),
            str(p.get("url_rating", "")),
            str(p.get("referring_domains", "")),
            str(p.get("backlinks", "")),
            str(p.get("organic_traffic", "")),
            p.get("content_richness_score", ""),
            p.get("diagnosis", ""),
            notes_str,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c == 12:  # Diagnosis column
                diag = p.get("diagnosis", "")
                if diag in DIAGNOSIS_FILLS:
                    cell.fill = DIAGNOSIS_FILLS[diag]
    ws3.freeze_panes = "B2"

    # --- Tab 4: Above The Fold — Mobile ---
    ws4 = wb.create_sheet("Above The Fold - Mobile")
    atf_headers = [
        "City", "Position", "URL", "Page Type", "Hero Headline", "Hero Subheadline",
        "Primary CTA Text", "CTA Location", "Has Trust Badge in Hero", "Has Video in Hero",
        "Has Background Image", "First Impression Summary",
    ]
    for c, h in enumerate(atf_headers, 1):
        ws4.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, p in enumerate(pages, 2):
        af = p.get("above_fold_mobile", {})
        summary = f"Headline: {af.get('headline', '')[:50]}..."
        row_data = [
            p.get("city", ""),
            p.get("position", ""),
            (p.get("url", "") or "")[:60],
            p.get("page_type", ""),
            af.get("headline", ""),
            af.get("subheadline", "")[:100] if af.get("subheadline") else "",
            af.get("cta_text", ""),
            "in hero" if af.get("cta_text") else "",
            "Yes" if af.get("has_trust_badge") else "No",
            "Yes" if af.get("has_video") else "No",
            "Yes" if af.get("has_background_image") else "No",
            summary,
        ]
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)
    ws4.freeze_panes = "B2"

    # --- Tab 5: Section Intelligence ---
    ws5 = wb.create_sheet("Section Intelligence")
    si_headers = [
        "Section Topic", "Pages Containing It", "Position 1 Pages That Include It",
        "Typical Position on Page", "Estimated Word Count Range", "Wireframe Recommendation",
    ]
    for c, h in enumerate(si_headers, 1):
        ws5.cell(row=1, column=c, value=h).font = Font(bold=True)
    si_key_map = {
        "Section Topic": "section_topic",
        "Pages Containing It": "pages_containing",
        "Position 1 Pages That Include It": "position_1_include",
        "Typical Position on Page": "typical_position",
        "Estimated Word Count Range": "word_count_range",
        "Wireframe Recommendation": "wireframe_recommendation",
    }
    for r, si in enumerate(analysis.get("section_intelligence", []), 2):
        for c, header in enumerate(si_headers, 1):
            k = si_key_map.get(header, header)
            ws5.cell(row=r, column=c, value=si.get(k, ""))
    ws5.freeze_panes = "B2"

    # --- Tab 6: Technology & Differentiation ---
    ws6 = wb.create_sheet("Technology & Differentiation")
    td_headers = [
        "City", "Position", "URL", "Page Type", "Technologies Mentioned",
        "Credential Claims", "Statistical Claims", "Financing Mentioned", "Unique Differentiators",
    ]
    for c, h in enumerate(td_headers, 1):
        ws6.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, p in enumerate(pages, 2):
        ce = p.get("content_elements", {})
        tech = ce.get("technology_names", {})
        tech_list = tech.get("found", []) if isinstance(tech, dict) else []
        cred = ce.get("surgeon_credentials", {})
        cred_text = cred.get("exact_text", "")[:200] if isinstance(cred, dict) else ""
        stats = ce.get("outcome_statistics", {})
        claims = stats.get("claims", []) if isinstance(stats, dict) else []
        fin = ce.get("financing", {})
        fin_yes = fin.get("present", False) if isinstance(fin, dict) else False
        row_data = [
            p.get("city", ""),
            p.get("position", ""),
            (p.get("url", "") or "")[:60],
            p.get("page_type", ""),
            ", ".join(tech_list) if tech_list else "",
            cred_text,
            " | ".join(claims[:3]) if claims else "",
            "Yes" if fin_yes else "No",
            "",
        ]
        for c, val in enumerate(row_data, 1):
            ws6.cell(row=r, column=c, value=val)
    ws6.freeze_panes = "B2"

    # --- Tab 7: Page Type Summary ---
    ws7 = wb.create_sheet("Page Type Summary")
    pt_headers = [
        "City", "Position", "URL", "Detected Page Type", "Classification Confidence",
        "URL Signal", "Content Signal", "Final Classification", "Wireframe Weight", "Notes",
    ]
    for c, h in enumerate(pt_headers, 1):
        ws7.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, p in enumerate(pages, 2):
        notes = []
        if p.get("js_rendering_flagged"):
            notes.append("⚠️ May require JS rendering")
        if p.get("page_type") == "Homepage" and config.get("flag_homepages"):
            notes.append("Homepage ranking — domain authority likely the primary ranking factor")
        row_data = [
            p.get("city", ""),
            p.get("position", ""),
            (p.get("url", "") or "")[:60],
            p.get("page_type", ""),
            p.get("page_type_confidence", ""),
            p.get("url_signal", p.get("page_type_prelim", "")),
            p.get("content_signal", ""),
            p.get("page_type", ""),
            p.get("wireframe_weight", ""),
            " ".join(notes),
        ]
        for c, val in enumerate(row_data, 1):
            ws7.cell(row=r, column=c, value=val)
    ws7.freeze_panes = "B2"

    # --- Tab 8: Raw Data ---
    ws8 = wb.create_sheet("Raw Data")
    raw_headers = [
        "city", "position", "url", "page_type", "page_title", "meta_description",
        "h1", "word_count", "domain_rating", "url_rating", "content_richness_score", "diagnosis",
        "js_rendering_flagged", "scrape_failed", "faq", "testimonials", "surgeon_credentials",
        "technology", "online_scheduling", "outcome_stats",
    ]
    for c, h in enumerate(raw_headers, 1):
        ws8.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, p in enumerate(pages, 2):
        ce = p.get("content_elements", {})
        struct = p.get("structure", {})
        row_data = [
            p.get("city", ""),
            p.get("position", ""),
            p.get("url", ""),
            p.get("page_type", ""),
            p.get("page_title", ""),
            p.get("meta_description", ""),
            struct.get("h1", ""),
            p.get("word_count", ""),
            str(p.get("domain_rating", "")),
            str(p.get("url_rating", "")),
            p.get("content_richness_score", ""),
            p.get("diagnosis", ""),
            "Yes" if p.get("js_rendering_flagged") else "No",
            "Yes" if p.get("scrape_failed") else "No",
            "Yes" if (ce.get("faq_section") or {}).get("present") else "No",
            "Yes" if (ce.get("testimonials") or {}).get("present") else "No",
            "Yes" if (ce.get("surgeon_credentials") or {}).get("present") else "No",
            ",".join((ce.get("technology_names") or {}).get("found", [])),
            "Yes" if (ce.get("online_scheduling") or {}).get("present") else "No",
            "Yes" if (ce.get("outcome_statistics") or {}).get("present") else "No",
        ]
        for c, val in enumerate(row_data, 1):
            ws8.cell(row=r, column=c, value=val)
    ws8.freeze_panes = "B2"

    wb.save(output_path)
    return output_path
